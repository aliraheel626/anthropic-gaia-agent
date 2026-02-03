"""File Handler Tool.

Provides multi-format file reading and parsing for GAIA benchmark tasks.
Supports PDF, images, spreadsheets, text files, and more.
"""

import base64
import io
import logging
import mimetypes
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Union

from claude_agent_sdk import tool

logger = logging.getLogger(__name__)


@dataclass
class FileContent:
    """Parsed file content."""

    filename: str
    mime_type: str
    text_content: Optional[str] = None
    binary_content: Optional[bytes] = None
    metadata: dict = None

    def __post_init__(self):
        if self.metadata is None:
            self.metadata = {}


class FileHandler:
    """Multi-format file handler for GAIA tasks."""

    SUPPORTED_EXTENSIONS = {
        # Text formats
        ".txt": "text/plain",
        ".md": "text/markdown",
        ".csv": "text/csv",
        ".json": "application/json",
        ".xml": "application/xml",
        ".html": "text/html",
        ".htm": "text/html",
        # Documents
        ".pdf": "application/pdf",
        # Spreadsheets
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        # Images
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
        # Code
        ".py": "text/x-python",
        ".js": "text/javascript",
        ".ts": "text/typescript",
        ".java": "text/x-java",
        ".cpp": "text/x-c++",
        ".c": "text/x-c",
    }

    def __init__(self):
        """Initialize file handler."""
        pass

    def read(self, path: Union[str, Path]) -> FileContent:
        """Read and parse a file.

        Args:
            path: Path to the file

        Returns:
            FileContent with parsed data
        """
        path = Path(path)

        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        ext = path.suffix.lower()
        mime_type = (
            self.SUPPORTED_EXTENSIONS.get(ext)
            or mimetypes.guess_type(str(path))[0]
            or "application/octet-stream"
        )

        # Route to appropriate parser
        if ext == ".pdf":
            return self._read_pdf(path, mime_type)
        elif ext in (".xlsx", ".xls"):
            return self._read_spreadsheet(path, mime_type)
        elif ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            return self._read_image(path, mime_type)
        elif ext == ".csv":
            return self._read_csv(path, mime_type)
        elif ext == ".json":
            return self._read_json(path, mime_type)
        else:
            return self._read_text(path, mime_type)

    def _read_text(self, path: Path, mime_type: str) -> FileContent:
        """Read a text file."""
        try:
            content = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            content = path.read_text(encoding="latin-1")

        return FileContent(
            filename=path.name,
            mime_type=mime_type,
            text_content=content,
            metadata={"size": path.stat().st_size},
        )

    def _read_pdf(self, path: Path, mime_type: str) -> FileContent:
        """Read a PDF file."""
        try:
            from pypdf import PdfReader

            reader = PdfReader(path)
            text_parts = []

            for i, page in enumerate(reader.pages):
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(f"--- Page {i + 1} ---\n{page_text}")

            return FileContent(
                filename=path.name,
                mime_type=mime_type,
                text_content="\n\n".join(text_parts),
                metadata={
                    "pages": len(reader.pages),
                    "size": path.stat().st_size,
                },
            )

        except ImportError:
            logger.error("pypdf not installed. Run: pip install pypdf")
            return FileContent(
                filename=path.name,
                mime_type=mime_type,
                binary_content=path.read_bytes(),
                metadata={"error": "pypdf not available"},
            )

    def _read_spreadsheet(self, path: Path, mime_type: str) -> FileContent:
        """Read an Excel spreadsheet."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, data_only=True)
            text_parts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"=== Sheet: {sheet_name} ===")

                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    rows.append("\t".join(row_values))

                text_parts.append("\n".join(rows))

            return FileContent(
                filename=path.name,
                mime_type=mime_type,
                text_content="\n\n".join(text_parts),
                metadata={
                    "sheets": wb.sheetnames,
                    "size": path.stat().st_size,
                },
            )

        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return FileContent(
                filename=path.name,
                mime_type=mime_type,
                binary_content=path.read_bytes(),
                metadata={"error": "openpyxl not available"},
            )

    def _read_csv(self, path: Path, mime_type: str) -> FileContent:
        """Read a CSV file."""
        import csv

        try:
            with open(path, "r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)

            # Format as text
            text_content = "\n".join(["\t".join(row) for row in rows])

            return FileContent(
                filename=path.name,
                mime_type=mime_type,
                text_content=text_content,
                metadata={
                    "rows": len(rows),
                    "columns": len(rows[0]) if rows else 0,
                    "size": path.stat().st_size,
                },
            )

        except Exception as e:
            logger.error(f"CSV read error: {e}")
            return self._read_text(path, mime_type)

    def _read_json(self, path: Path, mime_type: str) -> FileContent:
        """Read a JSON file."""
        import json

        try:
            content = path.read_text(encoding="utf-8")
            data = json.loads(content)

            # Pretty-print for readability
            formatted = json.dumps(data, indent=2)

            return FileContent(
                filename=path.name,
                mime_type=mime_type,
                text_content=formatted,
                metadata={"size": path.stat().st_size},
            )

        except json.JSONDecodeError as e:
            logger.warning(f"JSON parse error: {e}")
            return self._read_text(path, mime_type)

    def _read_image(self, path: Path, mime_type: str) -> FileContent:
        """Read an image file."""
        try:
            from PIL import Image

            with Image.open(path) as img:
                metadata = {
                    "width": img.width,
                    "height": img.height,
                    "format": img.format,
                    "mode": img.mode,
                    "size": path.stat().st_size,
                }

            # Read binary content
            binary_content = path.read_bytes()

            # For text extraction from images (OCR would go here)
            # For now, just provide metadata
            text_content = f"Image: {path.name}\nDimensions: {metadata['width']}x{metadata['height']}\nFormat: {metadata['format']}"

            return FileContent(
                filename=path.name,
                mime_type=mime_type,
                text_content=text_content,
                binary_content=binary_content,
                metadata=metadata,
            )

        except ImportError:
            logger.error("Pillow not installed. Run: pip install Pillow")
            return FileContent(
                filename=path.name,
                mime_type=mime_type,
                binary_content=path.read_bytes(),
                metadata={"error": "Pillow not available"},
            )

    def get_base64(self, path: Union[str, Path]) -> str:
        """Get base64-encoded content of a file.

        Args:
            path: Path to the file

        Returns:
            Base64-encoded string
        """
        path = Path(path)
        content = path.read_bytes()
        return base64.b64encode(content).decode("utf-8")


# Global file handler instance
_file_handler: Optional[FileHandler] = None


def get_file_handler() -> FileHandler:
    """Get or create the global file handler instance."""
    global _file_handler
    if _file_handler is None:
        _file_handler = FileHandler()
    return _file_handler


# Tool function for Claude Agent SDK
@tool(
    "read_file",
    "Read and parse a file. Supports text, PDF, spreadsheets (xlsx/xls), CSV, JSON, and images. Returns the extracted content as text.",
    {
        "path": str,
    },
)
async def read_file(args: dict) -> dict:
    """Read and parse a file.

    Args:
        args: Dictionary with 'path' (required)

    Returns:
        Dictionary with file content
    """
    file_path = args.get("path", "")

    if not file_path:
        return {"content": [{"type": "text", "text": "Error: No file path provided."}]}

    try:
        handler = get_file_handler()
        result = handler.read(file_path)

        response_parts = [f"**File: {result.filename}**"]
        response_parts.append(f"Type: {result.mime_type}")

        if result.metadata:
            meta_str = ", ".join(f"{k}: {v}" for k, v in result.metadata.items() if k != "error")
            if meta_str:
                response_parts.append(f"Metadata: {meta_str}")

        if result.text_content:
            # Truncate if too long
            content = result.text_content
            if len(content) > 15000:
                content = content[:15000] + "\n\n[Content truncated...]"
            response_parts.append(f"\n**Content:**\n{content}")
        elif result.binary_content:
            response_parts.append(f"\n(Binary content: {len(result.binary_content)} bytes)")

        return {"content": [{"type": "text", "text": "\n".join(response_parts)}]}

    except FileNotFoundError as e:
        return {"content": [{"type": "text", "text": f"File not found: {file_path}"}]}
    except Exception as e:
        logger.error(f"File read error: {e}")
        return {"content": [{"type": "text", "text": f"Error reading file: {str(e)}"}]}


@tool(
    "list_files",
    "List files in a directory. Returns filenames and basic info.",
    {
        "directory": str,
        "pattern": str,  # Optional glob pattern
    },
)
async def list_files(args: dict) -> dict:
    """List files in a directory.

    Args:
        args: Dictionary with 'directory' (required) and 'pattern' (optional)

    Returns:
        Dictionary with file list
    """
    directory = args.get("directory", "")
    pattern = args.get("pattern", "*")

    if not directory:
        return {"content": [{"type": "text", "text": "Error: No directory provided."}]}

    try:
        dir_path = Path(directory)

        if not dir_path.exists():
            return {"content": [{"type": "text", "text": f"Directory not found: {directory}"}]}

        if not dir_path.is_dir():
            return {"content": [{"type": "text", "text": f"Not a directory: {directory}"}]}

        files = list(dir_path.glob(pattern))

        if not files:
            return {
                "content": [
                    {"type": "text", "text": f"No files matching '{pattern}' in {directory}"}
                ]
            }

        response_parts = [f"**Files in {directory}** (pattern: {pattern}):\n"]

        for f in sorted(files)[:100]:  # Limit to 100 files
            if f.is_file():
                size = f.stat().st_size
                response_parts.append(f"- {f.name} ({size} bytes)")
            else:
                response_parts.append(f"- {f.name}/ (directory)")

        if len(files) > 100:
            response_parts.append(f"\n... and {len(files) - 100} more files")

        return {"content": [{"type": "text", "text": "\n".join(response_parts)}]}

    except Exception as e:
        logger.error(f"List files error: {e}")
        return {"content": [{"type": "text", "text": f"Error listing files: {str(e)}"}]}
